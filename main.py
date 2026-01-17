import akshare as ak
import pandas as pd
import numpy as np
from ta.trend import ADXIndicator
from ta.volume import OnBalanceVolumeIndicator, ChaikinMoneyFlowIndicator
from ta.volatility import BollingerBands
from datetime import datetime, timedelta, date
import os
import time
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import concurrent.futures
import random
import warnings

warnings.filterwarnings('ignore')

# --- 1. 全局配置 ---
CONFIG = {
    "IS_TEST_MODE": True,     # 测试模式(忽略周末/节假日，不写历史记录)
    "MIN_AMOUNT": 20000000,   # 2000万
    "MIN_PRICE": 2.5,         # 2.5元
    "MAX_WORKERS": 8,         # 线程数
    "DAYS_LOOKBACK": 250,     
    "BLACKLIST_DAYS": 30      
}

HISTORY_FILE = "stock_history_log.csv"
HOT_CONCEPTS = [] 
RESTRICTED_LIST = [] 
NORTHBOUND_SET = set() 
MARKET_ENV_TEXT = "⏳初始化..."

# --- 2. 交易日检查 ---
def is_trading_day():
    today = date.today()
    if today.weekday() >= 5: return False
    try:
        df_cal = ak.tool_trade_date_hist_sina()
        trade_dates = pd.to_datetime(df_cal['trade_date']).dt.date.tolist()
        return today in trade_dates
    except: return True

# --- 3. 市场情报 ---
def get_market_context():
    global HOT_CONCEPTS, RESTRICTED_LIST, NORTHBOUND_SET, MARKET_ENV_TEXT
    print("📡 [1/4] 连接交易所数据中心...")
    try:
        next_month = (datetime.now() + timedelta(days=CONFIG["BLACKLIST_DAYS"])).strftime("%Y-%m-%d")
        today = datetime.now().strftime("%Y-%m-%d")
        df_res = ak.stock_restricted_release_queue_em()
        cols = df_res.columns.tolist()
        code_col = next((c for c in cols if 'code' in c or '代码' in c), None)
        date_col = next((c for c in cols if 'date' in c or '时间' in c), None)
        if code_col and date_col:
            df_future = df_res[(df_res[date_col] >= today) & (df_res[date_col] <= next_month)]
            RESTRICTED_LIST = df_future[code_col].astype(str).tolist()
            print(f"🛡️ 已拉黑 {len(RESTRICTED_LIST)} 只解禁风险股")
    except: pass

    try:
        df = ak.stock_board_concept_name_em()
        df = df.sort_values(by="涨跌幅", ascending=False).head(15)
        HOT_CONCEPTS = df["板块名称"].tolist()
    except: pass

    try:
        df_sh = ak.stock_hsgt_top_10_em(symbol="沪股通")
        df_sz = ak.stock_hsgt_top_10_em(symbol="深股通")
        if df_sh is not None: NORTHBOUND_SET.update(df_sh['代码'].astype(str).tolist())
        if df_sz is not None: NORTHBOUND_SET.update(df_sz['代码'].astype(str).tolist())
        print(f"💰 北向重仓: {len(NORTHBOUND_SET)} 只")
    except: pass
    
    try:
        sh = ak.stock_zh_index_daily(symbol="sh000001")
        curr = sh.iloc[-1]
        ma20 = sh['close'].rolling(20).mean().iloc[-1]
        pct = (curr['close'] - sh.iloc[-2]['close']) / sh.iloc[-2]['close'] * 100
        status = "⛈️暴跌" if pct < -1.5 else ("🌧️空头" if curr['close'] < ma20 else "🌤️多头")
        MARKET_ENV_TEXT = f"上证: {curr['close']:.2f} ({pct:+.2f}%) | {status}"
        print(f"🌍 {MARKET_ENV_TEXT}")
    except: pass

def get_targets_robust():
    print(">>> [2/4] 全市场扫描与初筛...")
    try:
        df = ak.stock_zh_a_spot_em()
        col_map = {"最新价": "price", "成交额": "amount", "代码": "code", "名称": "name", 
                   "换手率": "turnover", "市盈率-动态": "pe", "市净率": "pb", "总市值": "mktcap"}
        df.rename(columns=col_map, inplace=True)
        for c in ["price", "amount", "turnover", "pe", "pb", "mktcap"]:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        
        df.dropna(subset=["price", "amount"], inplace=True)
        df = df[df["code"].str.startswith(("60", "00"))]
        df = df[~df['name'].str.contains('ST|退')]
        df = df[df["price"] >= CONFIG["MIN_PRICE"]]
        df = df[df["amount"] > CONFIG["MIN_AMOUNT"]]
        
        # 🔥 [修正] 移除 PB 限制，防止漏掉高估值妖股
        # 🔥 [修正] 移除 换手率限制，防止漏掉缩量好股
        # df = df[df["turnover"] >= 1.0] 
        # df = df[df["pb"] <= 20] 
        
        df = df[~df["code"].isin(RESTRICTED_LIST)]
        print(f"✅ 有效标的: {len(df)} 只 (已放宽初筛)")
        return df.to_dict('records')
    except: return []

def get_data_with_retry(code, start_date):
    time.sleep(random.uniform(0.01, 0.05))
    for _ in range(3):
        try:
            df = ak.stock_zh_a_hist(symbol=code, period="daily", start_date=start_date, adjust="qfq", timeout=5)
            if df is not None and not df.empty: return df
        except: time.sleep(0.2)
    return None

def get_60m_data_optimized(code):
    time.sleep(random.uniform(0.1, 0.4))
    for attempt in range(3):
        try:
            try:
                df = ak.stock_zh_a_hist_min_em(symbol=code, period="60", adjust="qfq", timeout=10)
            except:
                df = ak.stock_zh_a_hist_min_em(symbol=code, period="60", adjust="", timeout=10)
            if df is not None and not df.empty:
                df.rename(columns={"时间":"date","开盘":"open","收盘":"close","最高":"high","最低":"low","成交量":"volume"}, inplace=True)
                return df.tail(60) 
        except: time.sleep(1) 
    return None

def get_stock_catalysts(code):
    try:
        news_df = ak.stock_news_em(symbol=code)
        if not news_df.empty: return news_df.iloc[0]['新闻标题']
    except: pass
    return ""

def analyze_kline_health(df_full):
    if len(df_full) < 60: return "⚪数据不足", 0
    curr = df_full.iloc[-1]
    body_top = max(curr['open'], curr['close'])
    price_range = curr['high'] - curr['low']
    if price_range == 0: return "⚪极小波动", 0
    
    upper_ratio = (curr['high'] - body_top) / price_range
    vol_ratio = curr['volume'] / df_full['volume'].tail(5).mean()
    trend_up = curr['close'] > df_full['close'].tail(20).mean()

    if upper_ratio > 0.4:
        if vol_ratio > 2.0: return "⚠️高位抛压", -30
        elif not trend_up: return "📉冲高受阻", -10
        elif curr['close'] >= curr['open']: return "☝️仙人指路", 15
    elif (min(curr['open'], curr['close']) - curr['low']) / price_range > 0.4:
        if curr['low'] <= df_full['close'].tail(20).mean(): return "🛡️金针探底", 20
        return "⚓底部承接", 15
    elif (curr['close'] - curr['open']) / price_range > 0.6:
        prev_open = df_full['open'].iloc[-2]
        if curr['close'] > prev_open: return "⚡阳包阴", 25
        return "💪实体强攻", 10
    return "⚪普通震荡", 0

# --- 4. 核心逻辑 (🔥 软性过滤版) ---
def process_stock_logic(df, stock_info):
    code = stock_info['code']
    name = stock_info['name']
    pe = stock_info.get('pe', 0)
    turnover = stock_info.get('turnover', 0)
    mktcap = stock_info.get('mktcap', 0)

    if len(df) < 120: return None
    
    rename_dict = {"日期":"date","开盘":"open","收盘":"close","最高":"high","最低":"low","成交量":"volume","成交额":"amount"}
    col_map = {k:v for k,v in rename_dict.items() if k in df.columns}
    df.rename(columns=col_map, inplace=True)
    
    close = df["close"]
    high = df["high"]
    low = df["low"]
    volume = df["volume"]
    df["vwap"] = df["amount"] / volume if "amount" in df.columns else (high + low + close) / 3

    df["pct_chg"] = close.pct_change() * 100
    today_pct = df["pct_chg"].iloc[-1]
    
    try:
        last_3_pct = df["pct_chg"].tail(3).values
        if len(last_3_pct) == 3:
            pct_3day_str = f"{last_3_pct[0]:+.1f}% ➡ {last_3_pct[1]:+.1f}% ➡ {last_3_pct[2]:+.1f}%"
        else: pct_3day_str = "-"
    except: pct_3day_str = "-"
    
    df["MA5"] = close.rolling(5).mean()
    df["MA20"] = close.rolling(20).mean()
    df["MA60"] = close.rolling(60).mean()
    df["BIAS20"] = (close - df["MA20"]) / df["MA20"] * 100

    bb_ind = BollingerBands(close, window=20, window_dev=2)
    df["BB_Upper"] = bb_ind.bollinger_hband()
    df["BB_Lower"] = bb_ind.bollinger_lband()
    df["BB_Width"] = bb_ind.bollinger_wband()
    df["BB_PctB"] = bb_ind.bollinger_pband()

    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    df["DIF"] = ema12 - ema26
    df["DEA"] = df["DIF"].ewm(span=9, adjust=False).mean()
    df["MACD_Bar"] = (df["DIF"] - df["DEA"]) * 2
    
    low_9 = low.rolling(9, min_periods=9).min()
    high_9 = high.rolling(9, min_periods=9).max()
    rsv = (close - low_9) / (high_9 - low_9) * 100
    rsv = rsv.fillna(50)
    df['K'] = rsv.ewm(com=2, adjust=False).mean()
    df['D'] = df['K'].ewm(com=2, adjust=False).mean()
    df['J'] = 3 * df['K'] - 2 * df['D']
    
    delta = close.diff()
    up = delta.clip(lower=0)
    down = -1 * delta.clip(upper=0)
    ema_up = up.ewm(com=5, adjust=False).mean()
    ema_down = down.ewm(com=5, adjust=False).mean()
    rs = ema_up / ema_down
    df['RSI'] = 100 - (100 / (1 + rs))
    
    obv_ind = OnBalanceVolumeIndicator(close, volume)
    df["OBV"] = obv_ind.on_balance_volume()
    df["OBV_MA10"] = df["OBV"].rolling(10).mean()
    
    cmf_ind = ChaikinMoneyFlowIndicator(high, low, close, volume, window=20)
    df["CMF"] = cmf_ind.chaikin_money_flow()
    df["ADX"] = ADXIndicator(high, low, close, window=14).adx()

    curr = df.iloc[-1]
    prev = df.iloc[-2]
    
    has_zt = (df["pct_chg"].tail(30) > 9.5).sum() >= 1
    is_today_limit = curr["close"] >= round(prev["close"] * 1.095, 2)
    
    # 🔥 [修正] 将所有硬性过滤改为“软通过”，只剔除绝对垃圾股
    if turnover > 25 and not is_today_limit: return None
    if curr["J"] > 105: return None 
    
    # 注释掉严格过滤，确保有股入选，后续在评分里扣分即可
    # if curr["OBV"] <= curr["OBV_MA10"]: return None 
    # if curr["CMF"] < 0.05: return None
    # if curr["MACD_Bar"] <= prev["MACD_Bar"]: return None

    # 策略
    signal_type = ""
    suggest_buy = curr["close"]
    stop_loss = curr["MA20"]
    
    is_deep_dip = (prev["BIAS20"] < -8) or (prev["RSI"] < 20)
    is_reversal = (curr["close"] > curr["MA5"]) and (curr["pct_chg"] > 1.5)
    if is_deep_dip and is_reversal:
        signal_type = "⚱️黄金坑"; stop_loss = round(curr["low"] * 0.98, 2)
    
    if not signal_type and has_zt and curr["close"] > curr["MA60"]:
        vol_ratio = curr["volume"] / df["volume"].tail(5).mean()
        if vol_ratio < 0.8: # 放宽
            signal_type = "🐉龙回头"; stop_loss = round(df["BB_Lower"].iloc[-1], 2)
    
    if not signal_type and curr["close"] > curr["MA60"] and curr["CMF"] > 0.05 and curr["ADX"] > 20: # 放宽ADX
        signal_type = "🏦机构控盘"; suggest_buy = round(curr["vwap"], 2)
    
    if not signal_type and curr["close"] < curr["MA60"] * 1.2 and curr["BB_Width"] < 12:
        signal_type = "⚡底部变盘"

    chip_signal = ""
    if (curr["close"] - df["low"].tail(120).min()) / (df["high"].tail(120).max() - df["low"].tail(120).min() + 0.001) < 0.4:
        if df["close"].tail(60).std() / df["close"].tail(60).mean() < 0.15: chip_signal = "🏆筹码密集" 

    patterns = []
    if df[df['close']>df['open']].tail(20)['volume'].sum() > df[df['close']<df['open']].tail(20)['volume'].sum() * 2.0: patterns.append("🟥红肥绿瘦")
    if (prev['close'] < prev['open']) and (curr['close'] > curr['open']) and (curr['close'] > prev['open']): patterns.append("⚡N字反包")
    recent_5 = df.tail(5)
    if (recent_5['close'] > recent_5['MA5']).all() and (recent_5['pct_chg'].abs() < 4.0).all() and (recent_5['close'].iloc[-1] > recent_5['close'].iloc[0]):
        patterns.append("🐜蚂蚁上树")
    pattern_str = " ".join(patterns)
    
    is_macd_gold = (prev["DIF"] < prev["DEA"]) and (curr["DIF"] > curr["DEA"])
    is_kdj_gold = (prev["J"] < prev["K"]) and (curr["J"] > curr["K"]) and (curr["J"] < 80)
    
    # 🔥 [放宽入选] 只要有一点亮点就入选
    is_qualified = False
    if signal_type: is_qualified = True
    elif is_macd_gold or is_kdj_gold: is_qualified = True
    elif chip_signal and pattern_str: is_qualified = True
    elif code in NORTHBOUND_SET: is_qualified = True
    
    if not is_qualified: return None

    kline_status, kline_score = analyze_kline_health(df)

    status_60m = "⏳数据不足"
    try:
        df_60 = get_60m_data_optimized(code)
        if df_60 is not None and len(df_60) > 20:
            c60 = df_60["close"]
            m60 = c60.ewm(span=12, adjust=False).mean() - c60.ewm(span=26, adjust=False).mean()
            s60 = m60.ewm(span=9, adjust=False).mean()
            if m60.iloc[-2] < s60.iloc[-2] and m60.iloc[-1] > s60.iloc[-1]: status_60m = "✅60分金叉"
            elif m60.iloc[-1] > s60.iloc[-1]: status_60m = "🚀60分多头"
            else: status_60m = "⚠️60分回调"
        else: status_60m = "❌获取失败"
    except: status_60m = "🚫计算异常"

    cross_status = ""
    if is_macd_gold and is_kdj_gold: cross_status = "⚡双金叉"
    elif is_macd_gold: cross_status = "🔥MACD金叉"
    elif is_kdj_gold: cross_status = "📈KDJ金叉"
    elif signal_type == "⚱️黄金坑": cross_status = "🟢绿柱缩短"

    reasons = []
    if signal_type: reasons.append("策略")
    if chip_signal and pattern_str: reasons.append("筹/形共振")
    if cross_status == "⚡双金叉": reasons.append("双金叉")
    if code in NORTHBOUND_SET: reasons.append("外资重仓")
    resonance_str = "+".join(reasons)

    news_title = get_stock_catalysts(code)
    hot_matched = ""
    for hot in HOT_CONCEPTS:
        if hot in news_title: hot_matched = hot; break
    display_concept = f"🔥{hot_matched}" if hot_matched else ""

    macd_warn = "⛽空中加油" if (curr["DIF"]>curr["DEA"] and curr["DIF"]>0 and curr["MACD_Bar"]>prev["MACD_Bar"]) else ""
    bar_trend = "🔴红增" if curr["MACD_Bar"] > 0 else "🟢绿缩"
    final_macd = f"{bar_trend}|{macd_warn if macd_warn else cross_status}"
    bb_state = "🚀突破上轨" if curr["BB_PctB"] > 1.0 else ("↔️极度收口" if curr["BB_Width"] < 12 else "")

    reason_parts = []
    if signal_type: reason_parts.append(f"🎯{signal_type}")
    elif "金叉" in status_60m: reason_parts.append("✅60分金叉")
    elif is_macd_gold: reason_parts.append("📈MACD金叉")
    if hot_matched: reason_parts.append(f"🔥{hot_matched}")
    if code in NORTHBOUND_SET: reason_parts.append("💰北向")
    selection_reason = " + ".join(reason_parts) if reason_parts else "技术面好转"

    risk_list = []
    if pe < 0: risk_list.append("业绩亏损") 
    elif pe > 100: risk_list.append("估值过高")
    if curr["BIAS20"] > 15: risk_list.append("短期涨幅过大")
    if turnover > 15: risk_list.append("交易过热")
    if "回调" in status_60m: risk_list.append("短线有抛压")
    if curr["OBV"] <= curr["OBV_MA10"]: risk_list.append("资金流出") # 提示风险
    if curr["CMF"] < 0: risk_list.append("主力资金弱") # 提示风险
    if mktcap > 0 and mktcap/100000000 < 20: risk_list.append("微盘股波动大")
    risk_text = "；".join(risk_list) if risk_list else "暂无明显风险"

    advice_text = ""
    if signal_type == "⚱️黄金坑":
        advice_text = f"✅现价{curr['close']}可买，跌破{stop_loss}止损"
    elif signal_type == "🐉龙回头":
        advice_text = f"⏳建议在{suggest_buy}附近低吸，破{stop_loss}跑路"
    elif "金叉" in status_60m:
        advice_text = f"✅短线买点已现，现价{curr['close']}介入"
    else:
        advice_text = f"👀先观察，若站稳{suggest_buy}再买"

    # 多头排列检查
    is_bullish_trend = (curr['MA5'] > curr['MA10'] > curr['MA20'] > curr['MA60'])
    has_gap = curr['low'] > prev['high'] 

    return {
        "代码": code, "名称": name, 
        "选股理由": selection_reason, "风险提示": risk_text, "操作建议": advice_text,          
        "现价": curr["close"], "今日涨跌": f"{today_pct:+.2f}%", 
        "近3日涨跌": pct_3day_str,
        "K线形态": kline_status, "K线评分": kline_score,
        "60分状态": status_60m, "BIAS乖离": round(curr["BIAS20"], 1),
        "连续": "", "共振因子": resonance_str,
        "信号类型": signal_type, "热门概念": display_concept,
        "OBV状态": "🚀流入" if curr["OBV"] > curr["OBV_MA10"] else "⚠️流出",
        "筹码分布": chip_signal, "形态特征": pattern_str,
        "MACD状态": final_macd, "布林状态": bb_state,
        "今日CMF": round(curr["CMF"], 3), "昨日CMF": round(prev["CMF"], 3), "前日CMF": round(prev_2["CMF"], 3),
        "RSI指标": round(curr["RSI"], 1), "J值": round(curr["J"], 1),
        "建议挂单": suggest_buy, "止损价": stop_loss,
        "换手率": turnover, "市盈率": pe, "总市值": round(mktcap / 100000000, 2), 
        "有缺口": has_gap, "多头排列": is_bullish_trend
    }

def calculate_score_and_details(row):
    score = 0
    details = []
    
    trend_str = str(MARKET_ENV_TEXT)
    if "暴跌" in trend_str: score -= 50; details.append("⛈️大盘暴跌-50")
    elif "空头" in trend_str: score -= 15; details.append("🌧️大盘空头-15")
    elif "多头" in trend_str: score += 10; details.append("🌤️大盘多头+10")
    
    k_score = float(row.get('K线评分', 0))
    if k_score != 0: score += k_score; details.append(f"K线{k_score:+}")
    
    s60 = str(row.get('60分状态', ''))
    if "金叉" in s60: score += 100; details.append("✅60分金叉+100")
    elif "多头" in s60: score += 80; details.append("🚀60分多头+80")
    elif "回调" in s60: score -= 20; details.append("⚠️60分回调-20")
    
    streak = str(row.get('连续', ''))
    if "3连" in streak or "4连" in streak: score += 50; details.append("🔥连板+50")
    elif "2连" in streak: score += 30; details.append("🔥2连板+30")
    
    if row.get('有缺口', False): score += 20; details.append("🆙跳空缺口+20")
    if row.get('多头排列', False): score += 20; details.append("📈均线多头+20")

    try:
        c1, c2, c3 = float(row.get('今日CMF', 0)), float(row.get('昨日CMF', 0)), float(row.get('前日CMF', 0))
        if c1 > c2 > c3: score += 30; details.append("🔺资金加速+30")
        elif c1 > c2: score += 10; details.append("资金流入+10")
    except: pass
    
    if "外资" in str(row.get('共振因子', '')): score += 25; details.append("💰北向重仓+25")
    if "红肥" in str(row.get('形态特征', '')): score += 15; details.append("🟥红肥绿瘦+15")
    
    if "黄金坑" in str(row.get('信号类型', '')): score += 20; details.append("⚱️黄金坑+20")
    if "双金叉" in str(row.get('共振因子', '')): score += 15; details.append("⚡双金叉+15")
    if "🔥" in str(row.get('热门概念', '')): score += 15; details.append("🔥蹭热点+15")
    
    try:
        pe = float(row.get('市盈率', 0))
        if 0 < pe < 25: score += 25; details.append("💎绩优低估+25")
        elif 25 <= pe < 50: score += 10; details.append("⚖️估值合理+10")
        elif pe < 0: score -= 20; details.append("❌业绩亏损-20")
        elif pe > 150: score -= 15; details.append("🎈估值过高-15")
    except: pass
    
    try:
        mv = float(row.get('总市值', 0))
        if 30 < mv < 200: score += 15; details.append("🎯黄金市值+15")
        elif mv < 20: score -= 10; details.append("⚠️微盘股风险-10")
    except: pass
    
    try:
        bias = float(row.get('BIAS乖离', 0))
        if bias > 18: score -= 40; details.append("🚫乖离过大-40")
    except: pass
    
    # 扣分项 (原硬性过滤改扣分)
    if "流出" in str(row.get('OBV状态', '')): score -= 15; details.append("📉OBV流出-15")
    try:
        cmf_val = float(row.get('今日CMF', 0))
        if cmf_val < 0.05: score -= 10; details.append("💸资金弱-10")
    except: pass
    
    try:
        macd_str = str(row.get('MACD状态', ''))
        if "绿" in macd_str: score -= 10; details.append("💤动能弱-10")
    except: pass

    return score, " | ".join(details)

def update_history(current_results):
    today_str = datetime.now().strftime("%Y-%m-%d")
    try:
        if os.path.exists(HISTORY_FILE):
            hist_df = pd.read_csv(HISTORY_FILE)
            hist_df['date'] = hist_df['date'].astype(str)
        else: hist_df = pd.DataFrame(columns=["date", "code"])
    except: hist_df = pd.DataFrame(columns=["date", "code"])

    hist_df = hist_df[hist_df['date'] != today_str]
    sorted_dates = sorted(hist_df['date'].unique(), reverse=True)
    processed_results = []
    new_rows = []
    
    for res in current_results:
        code = res['code'] if 'code' in res else res['代码']
        streak = 1
        for d in sorted_dates:
            if not hist_df[(hist_df['date'] == d) & (hist_df['code'] == str(code))].empty: streak += 1
            else: break
        res['连续'] = f"🔥{streak}连" if streak >= 2 else "首榜"
        processed_results.append(res)
        new_rows.append({"date": today_str, "code": str(code)})

    if not CONFIG["IS_TEST_MODE"]:
        if new_rows:
            hist_df = pd.concat([hist_df, pd.DataFrame(new_rows)], ignore_index=True)
            hist_df.to_csv(HISTORY_FILE, index=False)
            print("💾 历史记录已更新 (生产模式)")
    else:
        print("🧪 测试模式：不更新历史记录文件")
        
    return processed_results

def save_and_beautify(data_list):
    dt_str = datetime.now().strftime("%Y%m%d_%H%M")
    mode_str = "测试" if CONFIG["IS_TEST_MODE"] else "实盘"
    filename = f"严选_绝对有票版_{mode_str}_{dt_str}.xlsx"
    
    if not data_list:
        pd.DataFrame([["无股入选 (请检查网络)"]]).to_excel(filename)
        return filename

    df = pd.DataFrame(data_list)
    res = df.apply(calculate_score_and_details, axis=1)
    df["综合评分"] = [x[0] for x in res]
    df["评分解析"] = [x[1] for x in res]
    
    cols = ["代码", "名称", "选股理由", "风险提示", "操作建议", "综合评分", "评分解析", 
            "现价", "今日涨跌", "近3日涨跌", 
            "总市值", "K线形态", "60分状态", 
            "BIAS乖离", "连续", "共振因子", "信号类型", "热门概念", "OBV状态", 
            "今日CMF", "昨日CMF", "前日CMF", "筹码分布", "形态特征", "MACD状态", 
            "布林状态", "RSI指标", "J值", "建议挂单", "止损价", "换手率", "市盈率"]
            
    for c in cols:
        if c not in df.columns: df[c] = ""
    df = df[cols]
    df.sort_values(by=["综合评分"], ascending=False, inplace=True)
    df.to_excel(filename, index=False)
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    fill_blue = PatternFill("solid", fgColor="4472C4")
    font_red = Font(color="FF0000", bold=True)
    font_green = Font(color="008000", bold=True)
    font_purple = Font(color="800080", bold=True)
    fill_yellow = PatternFill("solid", fgColor="FFF2CC")
    
    for cell in ws[1]:
        cell.fill = fill_blue
        cell.font = header_font
    
    for row in ws.iter_rows(min_row=2):
        if float(row[5].value) >= 150: row[5].fill = PatternFill("solid", fgColor="FFC7CE") 
        
        row[2].alignment = Alignment(horizontal='left')
        row[2].font = Font(bold=True, color="0000FF") 
        
        row[3].alignment = Alignment(horizontal='left') 
        row[3].font = Font(color="FF0000") 
        
        row[4].alignment = Alignment(horizontal='left') 
        row[4].font = Font(color="008000", bold=True) 
        
        row[6].alignment = Alignment(horizontal='left')
        row[6].font = Font(size=9)
        row[9].alignment = Alignment(horizontal='center')
        row[9].font = Font(size=9) 

        for idx in [8]: 
            val = str(row[idx].value)
            if "+" in val: row[idx].font = font_red
            elif "-" in val: row[idx].font = font_green
        
        if "金叉" in str(row[12].value): row[12].font = font_red; row[12].fill = fill_yellow
        elif "回调" in str(row[12].value): row[12].font = font_green
        
        if "外资" in str(row[15].value): row[15].font = font_red; row[15].fill = fill_yellow

    ws.column_dimensions['C'].width = 35 
    ws.column_dimensions['D'].width = 25 
    ws.column_dimensions['E'].width = 30 
    ws.column_dimensions['G'].width = 40 
    ws.column_dimensions['J'].width = 28
    
    start_row = ws.max_row + 3
    
    env_cell = ws.cell(row=start_row, column=1, value=f"🚥 {MARKET_ENV_TEXT}")
    env_cell.font = Font(size=14, bold=True, color="FFFFFF")
    if "多头" in MARKET_ENV_TEXT: env_cell.fill = PatternFill("solid", fgColor="008000")
    else: env_cell.fill = PatternFill("solid", fgColor="FFA500")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=29)
    start_row += 2

    cat_font = Font(name='微软雅黑', size=12, bold=True, color="0000FF")
    text_font = Font(name='微软雅黑', size=10)
    
    ws.cell(row=start_row, column=1, value="⚔️ 小白使用指南 (绝对有票版)").font = cat_font
    start_row += 1
    indicators = [
        ("选股理由", "告诉你为什么选它。如果出现'60分金叉'，说明日内走势好。"),
        ("风险提示", "看到'业绩亏损'或'资金流出'要小心，这些都会扣分。"),
        ("操作建议", "傻瓜指令。请严格执行'止损价'！"),
        ("综合评分", "分越高越安全。分数低但入选的，通常是形态好但基本面/资金面有瑕疵。")
    ]
    for name, desc in indicators:
        ws.cell(row=start_row, column=1, value=name).font = Font(bold=True)
        ws.cell(row=start_row, column=2, value=desc).font = text_font
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=15)
        start_row += 1

    wb.save(filename)
    print(f"✅ 结果已保存: {filename}")
    return filename

def analyze_one_stock(stock_info, start_dt):
    try:
        df = get_data_with_retry(stock_info['code'], start_dt)
        if df is None: return None
        return process_stock_logic(df, stock_info)
    except: return None

def main():
    print("=== A股严选 (最终绝对有票版: 软性过滤+保底输出) ===")
    
    if not CONFIG["IS_TEST_MODE"]:
        if not is_trading_day():
            print("😴 休市时间，程序自动退出。")
            return

    get_market_context() 
    start_time = time.time()
    targets = get_targets_robust()
    if not targets: return

    start_dt = (datetime.now() - timedelta(days=CONFIG["DAYS_LOOKBACK"])).strftime("%Y%m%d")
    
    print(f"🚀 待扫描: {len(targets)} 只 | 启动 {CONFIG['MAX_WORKERS']} 线程...")
    results = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=CONFIG["MAX_WORKERS"]) as executor:
        future_to_stock = {executor.submit(analyze_one_stock, r, start_dt): r['code'] for r in targets}
        count = 0
        total = len(targets)
        for future in concurrent.futures.as_completed(future_to_stock):
            count += 1
            if count % 50 == 0: print(f"进度: {count}/{total} ...")
            try:
                res = future.result()
                if res:
                    print(f"  ★ 入选: {res['名称']} -> {res['选股理由']}")
                    results.append(res)
            except: pass

    if results: results = update_history(results)
    print(f"\n耗时: {int(time.time() - start_time)}秒 | 选中 {len(results)} 只")
    save_and_beautify(results)

if __name__ == "__main__":
    main()
